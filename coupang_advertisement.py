import pymysql
from dotenv import load_dotenv
import os
from openpyxl import load_workbook

load_dotenv()

CLOUD_DB_HOST = os.getenv("CLOUD_DB_HOST")
CLOUD_DB_USER = os.getenv("CLOUD_DB_USER")
CLOUD_DB_PASSWORD = os.getenv("CLOUD_DB_PASSWORD")
CLOUD_DB_NAME = os.getenv("CLOUD_DB_NAME")

conn = pymysql.connect(
    host=CLOUD_DB_HOST,
    user=CLOUD_DB_USER,
    password=CLOUD_DB_PASSWORD,
    database=CLOUD_DB_NAME,
    charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor,
)

FILE_PATH = "/Users/ohnaseong/Downloads/2025_11.xlsx"
WORK_BOOK = load_workbook(FILE_PATH)


def insert_coupang_advertisement_table(conn):

    sheet = WORK_BOOK["coupang_advertisement"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "type": record[1],
                "vat_type": record[2],
                "campaign_type": record[3],
                "campaign_id": record[4],
                "campaign_name": record[5],
                "billable_advertising_fees": record[6],
            }
            data.append(dic)
        sql = """
            INSERT INTO coupang_advertisement_undirty (
                date,
                type,
                vat_type,
                campaign_type,
                campaign_id,
                campaign_name,
                billable_advertising_fees
            ) VALUES (
                %(date)s,
                %(type)s,
                %(vat_type)s,
                %(campaign_type)s,
                %(campaign_id)s,
                %(campaign_name)s,
                %(billable_advertising_fees)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into coupang_advertisement_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


insert_coupang_advertisement_table(conn)
# 모든 데이터베이스 연결 종료
conn.close()
