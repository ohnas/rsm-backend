import pymysql
from dotenv import load_dotenv
import os
from openpyxl import load_workbook

load_dotenv()

CLOUD_DB_HOST = os.getenv("CLOUD_DB_HOST")
CLOUD_DB_USER = os.getenv("CLOUD_DB_USER")
CLOUD_DB_PASSWORD = os.getenv("CLOUD_DB_PASSWORD")
CLOUD_DB_NAME = os.getenv("CLOUD_DB_NAME")

conn = pymysql.connect(
    host=CLOUD_DB_HOST,
    user=CLOUD_DB_USER,
    password=CLOUD_DB_PASSWORD,
    database=CLOUD_DB_NAME,
    charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor,
)

FILE_PATH = "/Users/ohnaseong/Downloads/2024.xlsx"
WORK_BOOK = load_workbook(FILE_PATH)


def insert_naver_advertisement_table(conn):

    sheet = WORK_BOOK["naver_advertisement"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "campaign_id": record[1],
                "campaign_name": record[2],
                "campaign_type": record[3],
                "imp_cnt": record[4],
                "clk_cnt": record[5],
                "ctr": record[6],
                "cpc": record[7],
                "sales_amt": record[8],
                "ccnt": record[9],
                "crto": record[10],
                "conv_amt": record[11],
                "ror": record[12],
                "cp_conv": record[13],
                "pc_nx_avg_rnk": record[14],
                "mbl_nx_avg_rnk": record[15],
                "avg_rnk": record[16],
            }
            data.append(dic)
        sql = """
            INSERT INTO naver_advertisement_undirty (
                date,
                campaign_id,
                campaign_name,
                campaign_type,
                imp_cnt,
                clk_cnt,
                ctr,
                cpc,
                sales_amt,
                ccnt,
                crto,
                conv_amt,
                ror,
                cp_conv,
                pc_nx_avg_rnk,
                mbl_nx_avg_rnk,
                avg_rnk
            ) VALUES (
                %(date)s,
                %(campaign_id)s,
                %(campaign_name)s,
                %(campaign_type)s,
                %(imp_cnt)s,
                %(clk_cnt)s,
                %(ctr)s,
                %(cpc)s,
                %(sales_amt)s,
                %(ccnt)s,
                %(crto)s,
                %(conv_amt)s,
                %(ror)s,
                %(cp_conv)s,
                %(pc_nx_avg_rnk)s,
                %(mbl_nx_avg_rnk)s,
                %(avg_rnk)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into naver_advertisement_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


insert_naver_advertisement_table(conn)
# 모든 데이터베이스 연결 종료
conn.close()
