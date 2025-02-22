import pymysql
from dotenv import load_dotenv
import os
from openpyxl import load_workbook

load_dotenv()

CLOUD_DB_HOST = os.getenv("CLOUD_DB_HOST")
CLOUD_DB_USER = os.getenv("CLOUD_DB_USER")
CLOUD_DB_PASSWORD = os.getenv("CLOUD_DB_PASSWORD")
CLOUD_DB_NAME = os.getenv("CLOUD_DB_NAME")


conn = pymysql.connect(
    host=CLOUD_DB_HOST,
    user=CLOUD_DB_USER,
    password=CLOUD_DB_PASSWORD,
    database=CLOUD_DB_NAME,
    charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor,
)

FILE_PATH = "/Users/ohnaseong/Downloads/2025_01.xlsx"
WORK_BOOK = load_workbook(FILE_PATH)


def insert_coupang_order_table(conn):

    sheet = WORK_BOOK["coupang_order"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "order_no": record[0],
                "tax_type": (None if record[1] == "" else record[1]),
                "sales_price": record[2],
                "refund_quantity": record[3],
                "sales_amount": record[4],
                "seller_discount_coupon": (None if record[5] == "" else record[5]),
                "seller_discount_coupon_a": (None if record[6] == "" else record[6]),
                "seller_discount_coupon_b": (None if record[7] == "" else record[7]),
                "sales_commission": record[8],
                "retroactive_payment_of_service_usage_fee": (
                    None if record[9] == "" else record[9]
                ),
                "my_shop_fee_discount": (None if record[10] == "" else record[10]),
                "settlement_amount": record[11],
                "payment_date": record[12],
                "delivery_date": (None if record[13] == "" else record[13]),
                "purchase_confirmation_date": record[14],
                "cancel_date": (None if record[15] == "" else record[15]),
                "purchase_confirmation_type": (
                    None if record[16] == "" else record[16]
                ),
                "settlement_date": record[17],
                "mode_shipping": record[18],
            }
            data.append(dic)
        sql = """
            INSERT INTO coupang_order_undirty (
                order_no,
                tax_type,
                sales_price,
                refund_quantity,
                sales_amount,
                seller_discount_coupon,
                seller_discount_coupon_a,
                seller_discount_coupon_b,
                sales_commission,
                retroactive_payment_of_service_usage_fee,
                my_shop_fee_discount,
                settlement_amount,
                payment_date,
                delivery_date,
                purchase_confirmation_date,
                cancel_date,
                purchase_confirmation_type,
                settlement_date,
                mode_shipping
            ) VALUES (
                %(order_no)s,
                %(tax_type)s,
                %(sales_price)s,
                %(refund_quantity)s,
                %(sales_amount)s,
                %(seller_discount_coupon)s,
                %(seller_discount_coupon_a)s,
                %(seller_discount_coupon_b)s,
                %(sales_commission)s,
                %(retroactive_payment_of_service_usage_fee)s,
                %(my_shop_fee_discount)s,
                %(settlement_amount)s,
                %(payment_date)s,
                %(delivery_date)s,
                %(purchase_confirmation_date)s,
                %(cancel_date)s,
                %(purchase_confirmation_type)s,
                %(settlement_date)s,
                %(mode_shipping)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(f"Success: Inserted {len(data)} rows into coupang_order_undirty")
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_coupang_order_detail_table(conn):

    sheet = WORK_BOOK["coupang_order_detail"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "order_no": record[0],
                "product_id": record[1],
                "product_name": record[2],
                "option_id": record[3],
                "option_name": record[4],
                "sales_quantity": record[5],
                "payment_date": record[6],
                "delivery_date": (None if record[7] == "" else record[7]),
                "purchase_confirmation_date": record[8],
                "cancel_date": (None if record[9] == "" else record[9]),
                "purchase_confirmation_type": record[10],
                "settlement_date": record[11],
            }
            data.append(dic)
        sql = """
            INSERT INTO coupang_order_detail_undirty (
                order_no,
                product_id,
                product_name,
                option_id,
                option_name,
                sales_quantity,
                payment_date,
                delivery_date,
                purchase_confirmation_date,
                cancel_date,
                purchase_confirmation_type,
                settlement_date
            ) VALUES (
                %(order_no)s,
                %(product_id)s,
                %(product_name)s,
                %(option_id)s,
                %(option_name)s,
                %(sales_quantity)s,
                %(payment_date)s,
                %(delivery_date)s,
                %(purchase_confirmation_date)s,
                %(cancel_date)s,
                %(purchase_confirmation_type)s,
                %(settlement_date)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into coupang_order_detail_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


insert_coupang_order_table(conn)
insert_coupang_order_detail_table(conn)
# 모든 데이터베이스 연결 종료
conn.close()
