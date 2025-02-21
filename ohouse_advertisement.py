import pymysql
from dotenv import load_dotenv
import os
from openpyxl import load_workbook

load_dotenv()

CLOUD_DB_HOST = os.getenv("CLOUD_DB_HOST")
CLOUD_DB_USER = os.getenv("CLOUD_DB_USER")
CLOUD_DB_PASSWORD = os.getenv("CLOUD_DB_PASSWORD")
CLOUD_DB_NAME = os.getenv("CLOUD_DB_NAME")

conn = pymysql.connect(
    host=CLOUD_DB_HOST,
    user=CLOUD_DB_USER,
    password=CLOUD_DB_PASSWORD,
    database=CLOUD_DB_NAME,
    charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor,
)

FILE_PATH = "/Users/ohnaseong/Downloads/2025_01.xlsx"
WORK_BOOK = load_workbook(FILE_PATH)


def insert_naver_advertisement_table(conn):

    sheet = WORK_BOOK["ohouse_advertisement"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "campaign_id": record[1],
                "campaign_name": record[2],
                "campaign_type": record[3],
                "impressions": (None if record[4] == "-" else record[4]),
                "clicks": record[5],
                "paid_advertising_fee": record[6],
                "free_advertising_fee": record[7],
                "cpc": record[8],
                "vcpm": (None if record[9] == "-" else record[9]),
            }
            data.append(dic)
        sql = """
            INSERT INTO ohouse_advertisement_undirty (
                date,
                campaign_id,
                campaign_name,
                campaign_type,
                impressions,
                clicks,
                paid_advertising_fee,
                free_advertising_fee,
                cpc,
                vcpm
            ) VALUES (
                %(date)s,
                %(campaign_id)s,
                %(campaign_name)s,
                %(campaign_type)s,
                %(impressions)s,
                %(clicks)s,
                %(paid_advertising_fee)s,
                %(free_advertising_fee)s,
                %(cpc)s,
                %(vcpm)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into ohouse_advertisement_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


insert_naver_advertisement_table(conn)
# 모든 데이터베이스 연결 종료
conn.close()
