import pymysql
from dotenv import load_dotenv
import os
from openpyxl import load_workbook


load_dotenv()

CLOUD_DB_HOST = os.getenv("CLOUD_DB_HOST")
CLOUD_DB_USER = os.getenv("CLOUD_DB_USER")
CLOUD_DB_PASSWORD = os.getenv("CLOUD_DB_PASSWORD")
CLOUD_DB_NAME = os.getenv("CLOUD_DB_NAME")


conn = pymysql.connect(
    host=CLOUD_DB_HOST,
    user=CLOUD_DB_USER,
    password=CLOUD_DB_PASSWORD,
    database=CLOUD_DB_NAME,
    charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor,
)

FILE_PATH = "/Users/ohnaseong/Downloads/2025_12_05.xlsx"
WORK_BOOK = load_workbook(FILE_PATH)


def insert_rocket_growth_order_table(conn):

    sheet = WORK_BOOK["rocket_growth_order"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "order_no": record[1],
                "transaction_type": record[2],
                "category_id": record[3],
                "category_name": record[4],
                "tax_type": record[5],
                "product_id": record[6],
                "option_id": record[7],
                "sku_id": record[8],
                "product_name": record[9],
                "option_name": record[10],
                "price": record[11],
                "quantity": record[12],
                "amount": record[13],
                "coupang_support_discount": record[14],
                "sales_amount": record[15],
                "instant_discount_coupon": record[16],
                "download_coupon": record[17],
                "seller_discount_coupon": record[18],
                "settlement_amount": record[19],
                "sales_commission_rate": record[20],
                "discounted_sales_commission_rate": record[21],
                "sales_commission": record[22],
                "sales_commission_vat": record[23],
            }
            data.append(dic)

        sql = """
            INSERT INTO rocket_growth_order_undirty (
                date,
                order_no,
                transaction_type,
                category_id,
                category_name,
                tax_type,
                product_id,
                option_id,
                sku_id,
                product_name,
                option_name,
                price,
                quantity,
                amount,
                coupang_support_discount,
                sales_amount,
                instant_discount_coupon,
                download_coupon,
                seller_discount_coupon,
                settlement_amount,
                sales_commission_rate,
                discounted_sales_commission_rate,
                sales_commission,
                sales_commission_vat
            ) VALUES (
                %(date)s,
                %(order_no)s,
                %(transaction_type)s,
                %(category_id)s,
                %(category_name)s,
                %(tax_type)s,
                %(product_id)s,
                %(option_id)s,
                %(sku_id)s,
                %(product_name)s,
                %(option_name)s,
                %(price)s,
                %(quantity)s,
                %(amount)s,
                %(coupang_support_discount)s,
                %(sales_amount)s,
                %(instant_discount_coupon)s,
                %(download_coupon)s,
                %(seller_discount_coupon)s,
                %(settlement_amount)s,
                %(sales_commission_rate)s,
                %(discounted_sales_commission_rate)s,
                %(sales_commission)s,
                %(sales_commission_vat)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into rocket_growth_order_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_rocket_growth_warehousing_table(conn):

    sheet = WORK_BOOK["rocket_growth_warehousing"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "transaction_type": record[1],
                "order_no": record[2],
                "shipping_no": record[3],
                "order_date": record[4],
                "product_id": record[5],
                "option_id": record[6],
                "sku_id": record[7],
                "product_name": record[8],
                "option_name": record[9],
                "price": record[10],
                "size": record[11],
                "center": record[12],
                "quantity": record[13],
                "purchase_quantity": record[14],
                "option_type": record[15],
                "fee": record[16],
                "discount": record[17],
                "net_amount": record[18],
            }
            data.append(dic)

        sql = """
            INSERT INTO rocket_growth_warehousing_undirty (
                date,
                transaction_type,
                order_no,
                shipping_no,
                order_date,
                product_id,
                option_id,
                sku_id,
                product_name,
                option_name,
                price,
                size,
                center,
                quantity,
                purchase_quantity,
                option_type,
                fee,
                discount,
                net_amount
            ) VALUES (
                %(date)s,
                %(transaction_type)s,
                %(order_no)s,
                %(shipping_no)s,
                %(order_date)s,
                %(product_id)s,
                %(option_id)s,
                %(sku_id)s,
                %(product_name)s,
                %(option_name)s,
                %(price)s,
                %(size)s,
                %(center)s,
                %(quantity)s,
                %(purchase_quantity)s,
                %(option_type)s,
                %(fee)s,
                %(discount)s,
                %(net_amount)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into rocket_growth_warehousing_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_rocket_growth_shipping_table(conn):

    sheet = WORK_BOOK["rocket_growth_shipping"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "transaction_type": record[1],
                "order_no": record[2],
                "shipping_no": record[3],
                "order_date": record[4],
                "product_id": record[5],
                "option_id": record[6],
                "sku_id": record[7],
                "product_name": record[8],
                "option_name": record[9],
                "price": record[10],
                "size": record[11],
                "total_size": record[12],
                "center": record[13],
                "quantity": record[14],
                "fee": record[15],
                "discount": record[16],
                "discounted_fee": record[17],
                "additional_fee": record[18],
                "net_amount": record[19],
            }
            data.append(dic)

        sql = """
            INSERT INTO rocket_growth_shipping_undirty (
                date,
                transaction_type,
                order_no,
                shipping_no,
                order_date,
                product_id,
                option_id,
                sku_id,
                product_name,
                option_name,
                price,
                size,
                total_size,
                center,
                quantity,
                fee,
                discount,
                discounted_fee,
                additional_fee,
                net_amount
            ) VALUES (
                %(date)s,
                %(transaction_type)s,
                %(order_no)s,
                %(shipping_no)s,
                %(order_date)s,
                %(product_id)s,
                %(option_id)s,
                %(sku_id)s,
                %(product_name)s,
                %(option_name)s,
                %(price)s,
                %(size)s,
                %(total_size)s,
                %(center)s,
                %(quantity)s,
                %(fee)s,
                %(discount)s,
                %(discounted_fee)s,
                %(additional_fee)s,
                %(net_amount)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into rocket_growth_shipping_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_rocket_growth_storage_table(conn):

    sheet = WORK_BOOK["rocket_growth_storage"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "display_date": record[1],
                "storage_period": record[2],
                "product_id": record[3],
                "option_id": record[4],
                "sku_id": record[5],
                "product_name": record[6],
                "option_name": record[7],
                "category_id": record[8],
                "category_name": record[9],
                "stock_quantity": record[10],
                "cbm": record[11],
                "fee": record[12],
                "discount": record[13],
                "net_amount": record[14],
            }
            data.append(dic)

        sql = """
            INSERT INTO rocket_growth_storage_undirty (
                date,
                display_date,
                storage_period,
                product_id,
                option_id,
                sku_id,
                product_name,
                option_name,
                category_id,
                category_name,
                stock_quantity,
                cbm,
                fee,
                discount,
                net_amount
            ) VALUES (
                %(date)s,
                %(display_date)s,
                %(storage_period)s,
                %(product_id)s,
                %(option_id)s,
                %(sku_id)s,
                %(product_name)s,
                %(option_name)s,
                %(category_id)s,
                %(category_name)s,
                %(stock_quantity)s,
                %(cbm)s,
                %(fee)s,
                %(discount)s,
                %(net_amount)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into rocket_growth_storage_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_rocket_growth_return_table(conn):

    sheet = WORK_BOOK["rocket_growth_return"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "transaction_type": record[1],
                "order_no": record[2],
                "order_date": record[3],
                "product_id": record[4],
                "option_id": record[5],
                "sku_id": record[6],
                "product_name": record[7],
                "option_name": record[8],
                "category_id": record[9],
                "category_name": record[10],
                "return_quantity": record[11],
                "free_promotion_type": record[12],
                "quantity_to_be_billed": record[13],
                "price": record[14],
                "instant_discount_coupon": record[15],
                "download_coupon": record[16],
                "sales_amount": record[17],
                "total_size": record[18],
                "fee": record[19],
                "discount": record[20],
                "discounted_fee": record[21],
                "free_promotion_fee": record[22],
                "net_amount": record[23],
            }
            data.append(dic)

        sql = """
            INSERT INTO rocket_growth_return_undirty (
                date,
                transaction_type,
                order_no,
                order_date,
                product_id,
                option_id,
                sku_id,
                product_name,
                option_name,
                category_id,
                category_name,
                return_quantity,
                free_promotion_type,
                quantity_to_be_billed,
                price,
                instant_discount_coupon,
                download_coupon,
                sales_amount,
                total_size,
                fee,
                discount,
                discounted_fee,
                free_promotion_fee,
                net_amount
            ) VALUES (
                %(date)s,
                %(transaction_type)s,
                %(order_no)s,
                %(order_date)s,
                %(product_id)s,
                %(option_id)s,
                %(sku_id)s,
                %(product_name)s,
                %(option_name)s,
                %(category_id)s,
                %(category_name)s,
                %(return_quantity)s,
                %(free_promotion_type)s,
                %(quantity_to_be_billed)s,
                %(price)s,
                %(instant_discount_coupon)s,
                %(download_coupon)s,
                %(sales_amount)s,
                %(total_size)s,
                %(fee)s,
                %(discount)s,
                %(discounted_fee)s,
                %(free_promotion_fee)s,
                %(net_amount)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into rocket_growth_return_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_rocket_growth_restocking_table(conn):

    sheet = WORK_BOOK["rocket_growth_restocking"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "transaction_type": record[1],
                "order_no": record[2],
                "order_date": record[3],
                "check_no": record[4],
                "product_id": record[5],
                "option_id": record[6],
                "sku_id": record[7],
                "product_name": record[8],
                "option_name": record[9],
                "price": record[10],
                "instant_discount_coupon": record[11],
                "download_coupon": record[12],
                "sales_amount": record[13],
                "restocking_quantity": record[14],
                "free_promotion_type": record[15],
                "quantity_to_be_billed": record[16],
                "fee": record[17],
                "discount": record[18],
                "discounted_fee": record[19],
                "free_promotion_fee": record[20],
                "net_amount": record[21],
            }
            data.append(dic)

        sql = """
            INSERT INTO rocket_growth_restocking_undirty (
                date,
                transaction_type,
                order_no,
                order_date,
                check_no,
                product_id,
                option_id,
                sku_id,
                product_name,
                option_name,
                price,
                instant_discount_coupon,
                download_coupon,
                sales_amount,
                restocking_quantity,
                free_promotion_type,
                quantity_to_be_billed,
                fee,
                discount,
                discounted_fee,
                free_promotion_fee,
                net_amount
            ) VALUES (
                %(date)s,
                %(transaction_type)s,
                %(order_no)s,
                %(order_date)s,
                %(check_no)s,
                %(product_id)s,
                %(option_id)s,
                %(sku_id)s,
                %(product_name)s,
                %(option_name)s,
                %(price)s,
                %(instant_discount_coupon)s,
                %(download_coupon)s,
                %(sales_amount)s,
                %(restocking_quantity)s,
                %(free_promotion_type)s,
                %(quantity_to_be_billed)s,
                %(fee)s,
                %(discount)s,
                %(discounted_fee)s,
                %(free_promotion_fee)s,
                %(net_amount)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into rocket_growth_restocking_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_rocket_growth_export_table(conn):

    sheet = WORK_BOOK["rocket_growth_export"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "transaction_type": record[1],
                "export_type": record[2],
                "export_no": record[3],
                "export_date": record[4],
                "product_id": record[5],
                "option_id": record[6],
                "sku_id": record[7],
                "product_name": record[8],
                "option_name": record[9],
                "export_quantity": record[10],
                "free_promotion_type": record[11],
                "quantity_to_be_billed": record[12],
                "fee": record[13],
                "discount": record[14],
                "discounted_fee": record[15],
                "free_promotion_fee": record[16],
                "net_amount": record[17],
            }
            data.append(dic)

        sql = """
            INSERT INTO rocket_growth_export_undirty (
                date,
                transaction_type,
                export_type,
                export_no,
                export_date,
                product_id,
                option_id,
                sku_id,
                product_name,
                option_name,
                export_quantity,
                free_promotion_type,
                quantity_to_be_billed,
                fee,
                discount,
                discounted_fee,
                free_promotion_fee,
                net_amount
            ) VALUES (
                %(date)s,
                %(transaction_type)s,
                %(export_type)s,
                %(export_no)s,
                %(export_date)s,
                %(product_id)s,
                %(option_id)s,
                %(sku_id)s,
                %(product_name)s,
                %(option_name)s,
                %(export_quantity)s,
                %(free_promotion_type)s,
                %(quantity_to_be_billed)s,
                %(fee)s,
                %(discount)s,
                %(discounted_fee)s,
                %(free_promotion_fee)s,
                %(net_amount)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into rocket_growth_export_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_rocket_growth_compensation_table(conn):

    sheet = WORK_BOOK["rocket_growth_compensation"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "end_date": record[1],
                "order_no": record[2],
                "type": record[3],
                "kind": record[4],
                "option_id": record[5],
                "product_name": record[6],
                "option_name": record[7],
                "check_class": record[8],
                "check_no": record[9],
                "responsibility": record[10],
                "compensation_standard_amount": record[11],
                "compensation_rate": record[12],
                "compensation_amount": record[13],
            }
            data.append(dic)

        sql = """
            INSERT INTO rocket_growth_compensation_undirty (
                date,
                end_date,
                order_no,
                type,
                kind,
                option_id,
                product_name,
                option_name,
                check_class,
                check_no,
                responsibility,
                compensation_standard_amount,
                compensation_rate,
                compensation_amount
            ) VALUES (
                %(date)s,
                %(end_date)s,
                %(order_no)s,
                %(type)s,
                %(kind)s,
                %(option_id)s,
                %(product_name)s,
                %(option_name)s,
                %(check_class)s,
                %(check_no)s,
                %(responsibility)s,
                %(compensation_standard_amount)s,
                %(compensation_rate)s,
                %(compensation_amount)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into rocket_growth_compensation_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_rocket_growth_barcode_fee_table(conn):

    sheet = WORK_BOOK["rocket_growth_barcode_fee"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "transaction_type": record[1],
                "service_type": record[2],
                "inbound_no": record[3],
                "request_no": record[4],
                "product_id": record[5],
                "option_id": record[6],
                "sku_id": record[7],
                "product_name": record[8],
                "option_name": record[9],
                "size": record[10],
                "center": record[11],
                "quantity": record[12],
                "fee": record[13],
                "discount": record[14],
                "net_amount": record[15],
            }
            data.append(dic)

        sql = """
            INSERT INTO rocket_growth_barcode_fee_undirty (
                date,
                transaction_type,
                service_type,
                inbound_no,
                request_no,
                product_id,
                option_id,
                sku_id,
                product_name,
                option_name,
                size,
                center,
                quantity,
                fee,
                discount,
                net_amount
            ) VALUES (
                %(date)s,
                %(transaction_type)s,
                %(service_type)s,
                %(inbound_no)s,
                %(request_no)s,
                %(product_id)s,
                %(option_id)s,
                %(sku_id)s,
                %(product_name)s,
                %(option_name)s,
                %(size)s,
                %(center)s,
                %(quantity)s,
                %(fee)s,
                %(discount)s,
                %(net_amount)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(
                f"Success: Inserted {len(data)} rows into rocket_growth_barcode_fee_undirty"
            )
        except Exception as e:
            print(f"Error: {str(e)}")


insert_rocket_growth_order_table(conn)
insert_rocket_growth_warehousing_table(conn)
insert_rocket_growth_shipping_table(conn)
insert_rocket_growth_storage_table(conn)
insert_rocket_growth_return_table(conn)
insert_rocket_growth_restocking_table(conn)
insert_rocket_growth_export_table(conn)
insert_rocket_growth_compensation_table(conn)
insert_rocket_growth_barcode_fee_table(conn)

# 모든 데이터베이스 연결 종료
conn.close()
