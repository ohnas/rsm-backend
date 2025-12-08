import pymysql
from dotenv import load_dotenv
import os
from openpyxl import load_workbook

load_dotenv()

CLOUD_DB_HOST = os.getenv("CLOUD_DB_HOST")
CLOUD_DB_USER = os.getenv("CLOUD_DB_USER")
CLOUD_DB_PASSWORD = os.getenv("CLOUD_DB_PASSWORD")
CLOUD_DB_NAME = os.getenv("CLOUD_DB_NAME")

conn = pymysql.connect(
    host=CLOUD_DB_HOST,
    user=CLOUD_DB_USER,
    password=CLOUD_DB_PASSWORD,
    database=CLOUD_DB_NAME,
    charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor,
)

FILE_PATH = "/Users/ohnaseong/Downloads/2025_11.xlsx"
WORK_BOOK = load_workbook(FILE_PATH)


def insert_ohouse_order_table(conn):

    sheet = WORK_BOOK["ohouse"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "order_no": record[0],
                "order_prod_no": record[1],
                "order_option_no": record[2],
                "prod_id": record[3],
                "prod_my_code": (None if record[4] == "" else record[4]),
                "prod_name": record[5],
                "option_id": record[6],
                "option_my_code": record[7],
                "option_name": record[8],
                "quantity": record[9],
                "order_note": (None if record[10] == "" else record[10]),
                "is_today_departure": (None if record[11] == "" else record[11]),
                "is_assembly_request": (None if record[12] == "" else record[12]),
                "order_date": record[13],
                "expected_ship_date": record[14],
                "expected_delivery_date": record[15],
                "request_delivery_date": (None if record[16] == "" else record[16]),
                "fast_delivery_date": (None if record[17] == "" else record[17]),
                "purchase_confirmation_date": record[18],
                "personal_customs_code": (None if record[19] == "" else record[19]),
                "sales_price": record[20],
                "sales_price_by_quantity": record[21],
                "assembly_cost": record[22],
                "delivery_fee": record[23],
                "total_price": record[24],
                "settlement_amount": record[25],
                "is_island": (None if record[26] == "" else record[26]),
                "island_delivery_fee": record[27],
                "order_status": record[28],
                "mode_shipping": record[29],
            }
            data.append(dic)
        sql = """
            INSERT INTO ohouse_order_undirty (
                order_no,
                order_prod_no,
                order_option_no,
                prod_id,
                prod_my_code,
                prod_name,
                option_id,
                option_my_code,
                option_name,
                quantity,
                order_note,
                is_today_departure,
                is_assembly_request,
                order_date,
                expected_ship_date,
                expected_delivery_date,
                request_delivery_date,
                fast_delivery_date,
                purchase_confirmation_date,
                personal_customs_code,
                sales_price,
                sales_price_by_quantity,
                assembly_cost,
                delivery_fee,
                total_price,
                settlement_amount,
                is_island,
                island_delivery_fee,
                order_status,
                mode_shipping
            ) VALUES (
                %(order_no)s,
                %(order_prod_no)s,
                %(order_option_no)s,
                %(prod_id)s,
                %(prod_my_code)s,
                %(prod_name)s,
                %(option_id)s,
                %(option_my_code)s,
                %(option_name)s,
                %(quantity)s,
                %(order_note)s,
                %(is_today_departure)s,
                %(is_assembly_request)s,
                %(order_date)s,
                %(expected_ship_date)s,
                %(expected_delivery_date)s,
                %(request_delivery_date)s,
                %(fast_delivery_date)s,
                %(purchase_confirmation_date)s,
                %(personal_customs_code)s,
                %(sales_price)s,
                %(sales_price_by_quantity)s,
                %(assembly_cost)s,
                %(delivery_fee)s,
                %(total_price)s,
                %(settlement_amount)s,
                %(is_island)s,
                %(island_delivery_fee)s,
                %(order_status)s,
                %(mode_shipping)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(f"Success: Inserted {len(data)} rows into ohouse_order_undirty")
        except Exception as e:
            print(f"Error: {str(e)}")


insert_ohouse_order_table(conn)
# 모든 데이터베이스 연결 종료
conn.close()
