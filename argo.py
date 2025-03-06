import pymysql
from dotenv import load_dotenv
import os
from openpyxl import load_workbook


load_dotenv()

CLOUD_DB_HOST = os.getenv("CLOUD_DB_HOST")
CLOUD_DB_USER = os.getenv("CLOUD_DB_USER")
CLOUD_DB_PASSWORD = os.getenv("CLOUD_DB_PASSWORD")
CLOUD_DB_NAME = os.getenv("CLOUD_DB_NAME")


conn = pymysql.connect(
    host=CLOUD_DB_HOST,
    user=CLOUD_DB_USER,
    password=CLOUD_DB_PASSWORD,
    database=CLOUD_DB_NAME,
    charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor,
)

FILE_PATH = "/Users/ohnaseong/Downloads/2025_02.xlsx"
WORK_BOOK = load_workbook(FILE_PATH)


def insert_argo_inbound_table(conn):

    sheet = WORK_BOOK["argo_inbound"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "inbound_no": record[1],
                "net_amount": record[2],
                "brand": record[3],
            }
            data.append(dic)

        sql = """
            INSERT INTO argo_inbound (
                date, inbound_no, net_amount, brand
            ) VALUES (
                %(date)s, %(inbound_no)s, %(net_amount)s, %(brand)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(f"Success: Inserted {len(data)} rows into argo_inbound")
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_argo_return_table(conn):

    sheet = WORK_BOOK["argo_return"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "return_tracking_no": record[1],
                "net_amount": record[2],
                "brand": record[3],
            }
            data.append(dic)

        sql = """
            INSERT INTO argo_return (
                date, return_tracking_no, net_amount, brand
            ) VALUES (
                %(date)s, %(return_tracking_no)s, %(net_amount)s, %(brand)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(f"Success: Inserted {len(data)} rows into argo_return")
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_argo_storage_table(conn):

    sheet = WORK_BOOK["argo_storage"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []
    if records:
        for record in records:
            dic = {
                "date": record[0],
                "sku_name": record[1],
                "sku_barcode": record[2],
                "net_amount": record[3],
                "brand": record[4],
            }
            data.append(dic)

        sql = """
            INSERT INTO argo_storage (
                date, sku_name, sku_barcode, net_amount, brand
            ) VALUES (
                %(date)s, %(sku_name)s, %(sku_barcode)s, %(net_amount)s, %(brand)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(f"Success: Inserted {len(data)} rows into argo_storage")
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_argo_shipping_table(conn):

    sheet = WORK_BOOK["argo_shipping"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "order_no": record[1],
                "store_name": record[2],
                "channel": record[3],
                "tracking_no": record[4],
                "box_name": record[5],
                "size": record[6],
                "type": record[7],
                "kind": record[8],
                "sku_count": record[9],
                "net_amount": record[10],
                "brand": record[11],
            }
            data.append(dic)

        sql = """
            INSERT INTO argo_shipping (
                date,
                order_no,
                store_name,
                channel,
                tracking_no,
                box_name,
                size,
                type,
                kind,
                sku_count,
                net_amount,
                brand
            ) VALUES (
                %(date)s,
                %(order_no)s,
                %(store_name)s,
                %(channel)s,
                %(tracking_no)s,
                %(box_name)s,
                %(size)s,
                %(type)s,
                %(kind)s,
                %(sku_count)s,
                %(net_amount)s,
                %(brand)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(f"Success: Inserted {len(data)} rows into argo_shipping")
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_argo_b2b_table(conn):

    sheet = WORK_BOOK["argo_b2b"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "order_no": record[1],
                "tracking_no": record[2],
                "type": record[3],
                "location": record[4],
                "pallet_count": record[5],
                "box_count": record[6],
                "net_amount": record[7],
                "brand": record[8],
            }
            data.append(dic)

        sql = """
            INSERT INTO argo_b2b (
                date,
                order_no,
                tracking_no,
                type,
                location,
                pallet_count,
                box_count,
                net_amount,
                brand
            ) VALUES (
                %(date)s,
                %(order_no)s,
                %(tracking_no)s,
                %(type)s,
                %(location)s,
                %(pallet_count)s,
                %(box_count)s,
                %(net_amount)s,
                %(brand)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(f"Success: Inserted {len(data)} rows into argo_b2b")
        except Exception as e:
            print(f"Error: {str(e)}")


def insert_argo_etc_table(conn):

    sheet = WORK_BOOK["argo_etc"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "check_no": record[1],
                "reason": record[2],
                "kind": record[3],
                "count": record[4],
                "net_amount": record[5],
                "brand": record[6],
            }
            data.append(dic)

        sql = """
            INSERT INTO argo_etc (
                date,
                check_no,
                reason,
                kind,
                count,
                net_amount,
                brand
            ) VALUES (
                %(date)s,
                %(check_no)s,
                %(reason)s,
                %(kind)s,
                %(count)s,
                %(net_amount)s,
                %(brand)s
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(f"Success: Inserted {len(data)} rows into argo_etc")
        except Exception as e:
            print(f"Error: {str(e)}")


insert_argo_inbound_table(conn)
insert_argo_return_table(conn)
insert_argo_storage_table(conn)
insert_argo_shipping_table(conn)
insert_argo_b2b_table(conn)
insert_argo_etc_table(conn)

# 모든 데이터베이스 연결 종료
conn.close()
