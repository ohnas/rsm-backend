import pymysql
from dotenv import load_dotenv
import os
from openpyxl import load_workbook

load_dotenv()

CLOUD_DB_HOST = os.getenv("CLOUD_DB_HOST")
CLOUD_DB_USER = os.getenv("CLOUD_DB_USER")
CLOUD_DB_PASSWORD = os.getenv("CLOUD_DB_PASSWORD")
CLOUD_DB_NAME = os.getenv("CLOUD_DB_NAME")


conn = pymysql.connect(
    host=CLOUD_DB_HOST,
    user=CLOUD_DB_USER,
    password=CLOUD_DB_PASSWORD,
    database=CLOUD_DB_NAME,
    charset="utf8mb4",
    cursorclass=pymysql.cursors.DictCursor,
)

FILE_PATH = "/Users/ohnaseong/Downloads/2025_09.xlsx"
WORK_BOOK = load_workbook(FILE_PATH)


def insert_b2b_table(conn):

    sheet = WORK_BOOK["b2b"]
    records = list(sheet.iter_rows(values_only=True))[1:]
    data = []

    if records:
        for record in records:
            dic = {
                "date": record[0],
                "prod_name": record[1],
                "count": record[2],
                "price": record[3],
                "sales": record[4],
                "unit_cost": record[5],
                "cost": record[6],
                "type": record[7],
                "company_name": record[8],
                "comment": record[9],
                "brand": record[10],
            }
            data.append(dic)
        sql = """
            INSERT INTO b2b (
                date, prod_name, count, price, sales, unit_cost, cost, type, company_name, comment, brand
            ) VALUES (
                %(date)s, %(prod_name)s, %(count)s, %(price)s, %(sales)s, %(unit_cost)s, %(cost)s, %(type)s, %(company_name)s, %(comment)s, %(brand)s 
            )
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(sql, data)
            conn.commit()
            print(f"Success: Inserted {len(data)} rows into b2b")
        except Exception as e:
            print(f"Error: {str(e)}")


insert_b2b_table(conn)
# 모든 데이터베이스 연결 종료
conn.close()
